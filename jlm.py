import os
from asyncio.windows_events import NULL
import math
import openpyxl

import tkinter as tk
from tkinter import *
from tkinter import filedialog

from pygments.lexers.jvm import JavaLexer
from pygments import lex

#初回のみデータ蓄積用Excelファイルを作成
if(os.path.exists('javaLM.xlsx')):
    ()
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'language model 1'
    sheet['A1'] = 'token name'
    sheet['A2'] = 'int'
    sheet['B2'] = 1
    wb.save('javaLM.xlsx')

filepath = 'javaLM.xlsx'
wb = openpyxl.load_workbook(filename=filepath)
ws1 = wb['language model 1']

# ファイル指定の関数
def filedialog_clicked():
    fTyp = [("", "*")]
    iFile = os.path.abspath(os.path.dirname(__file__))
    iFilePath = filedialog.askopenfilename(filetype = fTyp, initialdir = iFile)
    entry2.set(iFilePath)

#言語モデルデータ追加関数
def add_data():

    filename= Entry_ref.get()
    fileobj = open(filename, "r", encoding="utf_8")

    while True:
        line = fileobj.readline()
        if line:
            words=lex(line, JavaLexer())
            for word in words:
                x_max_row = ws1.max_row
                for i in range(2,x_max_row+1):
                    if(str(word[1])==ws1.cell(row = i, column = 1).value):
                        ws1.cell(row = i, column = 2).value+=1
                        break
                    elif(i==x_max_row):
                        ws1.cell(row = i+1, column = 1).value=str(word[1])
                        ws1.cell(row = i+1, column = 2).value=1
        else:
            #言語モデルの統計にデータを追加
            x_max_row = ws1.max_row
            sum=0
            for i in range(2,x_max_row+1):
                sum+=ws1.cell(row = i, column = 2).value

            for i in range(2,x_max_row+1):
                ws1.cell(row = i, column = 3).value=round(math.log10(ws1.cell(row = i, column = 2).value/sum), 5)

            print(sum)
            ws1.cell(row = 1, column = 2).value=sum

            wb.save('javaLM.xlsx')
            break

# コード表示&自然さ計測関数
def display():

    filename= Entry_ref.get()
    basename = os.path.basename(filename)
    #Lnum=行数
    Lnum=1
    fileobj = open(filename, "r", encoding="utf_8")

    canvas.create_rectangle(930, 30, 1100, 3000, fill ='#f8f8ff')
    canvas.create_text(30, 10, text=basename,anchor="w",font=("", 20))
    canvas.create_text(950, 10, text="Score",anchor="w",font=("", 18))

    #最終行取得
    x_max_row = ws1.max_row
    #コードを表示する時に無理矢理文字数で間隔を開けてそれっぽく表示するために使用
    Cnum=0

    while True:
        line = fileobj.readline()
        if line:
            Score=0
            Lnum+=1

            words=lex(line, JavaLexer())
            Cnum=0
            for word in words:
                

                #スコア集計
                for i in range(2,x_max_row+1):
                    if(str(word[1])==ws1.cell(row = i, column = 1).value):
                        Score+=ws1.cell(row = i, column = 3).value
                        break
                    elif(i==x_max_row):
                        ()

                if str(word[0])=="Token.Name.Class":
                    Color="#800080"
                elif str(word[0])=="Token.Name":
                    Color='blue'
                elif str(word[0])=="Token.Punctuation":
                    Color="#008b8b"
                elif str(word[0])=="Token.Operator":
                    Color="#4169e1"
                else:
                    Color='black'
                
                canvas.create_text(60+Cnum*12, Lnum*22,text=word[1],anchor="w",fill=Color,font=("", 18))
                Cnum+=len(word[1])

            #行数,スコア表示,色分け
            if Score>=-5:
                Color="#0000ff"
            elif Score>=-15:
                Color='#0000cd'
            elif Score>=-25:
                Color="#000080"
            else:
                Color="#800080"

            canvas.create_text(20, Lnum*22, text=Lnum-1,font=("", 18))
            canvas.create_text(1000, Lnum*22, text=round(Score,5),fill=Color,font=("", 18))
            #canvas.create_line(0, Lnum*22+11, 1200, Lnum*22+11,width=0.1,fill='#f8f8ff')
      
        else:
            bar_y = tk.Scrollbar(canvas, orient=tk.VERTICAL)
            bar_x = tk.Scrollbar(canvas, orient=tk.HORIZONTAL)
            bar_y.pack(side=tk.RIGHT, fill=tk.Y)
            bar_x.pack(side=tk.BOTTOM, fill=tk.X)
            bar_y.config(command=canvas.yview)
            bar_x.config(command=canvas.xview)
            canvas.config(yscrollcommand=bar_y.set, xscrollcommand=bar_x.set)
            # Canvasのスクロール範囲を設定
            canvas.config(scrollregion=(0, 0, 1000, (Lnum+5)*22))
            break

if __name__ == "__main__":

    # rootの作成
    root = tk.Tk()
    root.title("JAVA LM")
    root.geometry("1200x1000+0+0")

    # 操作画面の作成
    frame2 = tk.Frame(root,height=200,pady=10,padx=20,relief=tk.RAISED, bd=4, bg='#4169e1')
    frame2.pack(fill=tk.X)

    # 「ファイル参照」ボタンの作成
    Button_ref = tk.Button(frame2, text="参照",font=("", 18), command=filedialog_clicked)
    Button_ref.pack(side=LEFT)

    # 「ファイル参照」エントリーの作成
    entry2 = StringVar()
    Entry_ref = tk.Entry(frame2, textvariable=entry2,width=40,font=("", 20))
    Entry_ref.pack(side=LEFT)

    # 「データ追加」ボタンの作成
    Button_ref = tk.Button(frame2, text="追加",font=("", 18), command=add_data)
    Button_ref.pack(side=RIGHT)

    # 「自然さ計測」ボタンの作成
    Button_ref = tk.Button(frame2, text="出力",font=("", 18), command=display)
    Button_ref.pack(side=RIGHT)

    #ソースコード出力画面
    canvas = tk.Canvas(root)
    canvas.place(x=0, y=75,width=1200, height=600)

    root.mainloop()